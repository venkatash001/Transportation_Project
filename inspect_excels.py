import openpyxl, os

files = [
    r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - CareOps_Centralized_Report\999. GCP\28.Transportation_Roster_Project\Sample Roster.xlsx",
    r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - CareOps_Centralized_Report\999. GCP\Skill Matrix.xlsx",
    r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - CareOps_Centralized_Report\999. GCP\TeamName Matrix.xlsx",
    r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - CareOps_Centralized_Report\999. GCP\LOB Worked Mapping.xlsx",
]

for path in files:
    if not os.path.exists(path):
        print(f"NOT FOUND: {os.path.basename(path)}")
        continue
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n{'='*60}")
    print(f"FILE: {os.path.basename(path)}")
    print(f"Sheets: {wb.sheetnames}")
    for sh_name in wb.sheetnames[:3]:
        ws = wb[sh_name]
        print(f"\n  Sheet: '{sh_name}'  ({ws.max_row} rows x {ws.max_column} cols)")
        # Print first 3 rows
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        for i, row in enumerate(rows, 1):
            clean = [str(c)[:25] if c is not None else '' for c in row[:12]]
            print(f"    Row {i}: {clean}")
    wb.close()
