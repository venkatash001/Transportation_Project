"""
routes/export.py - CSV, Excel, and PDF export endpoints.
"""
import sys
import io
import csv
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from datetime import date, timedelta
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from config import FUTURE_DAYS, HISTORICAL_DAYS
from db import get_roster_for_range, get_overrides_for_range
from matrix import build_matrix, date_range


def _lbl(cell) -> str:
    """Extract label from a cell (str or dict)."""
    return cell["label"] if isinstance(cell, dict) else str(cell)

router = APIRouter()

_IDENTITY_HEADERS = ["S.No", "User ID", "WIN ID", "Full Name", "Team", "Role", "L1 Manager"]


def _get_range(scope: str) -> tuple[date, date]:
    today = date.today()
    if scope == "historical":
        return today - timedelta(days=HISTORICAL_DAYS), today - timedelta(days=1)
    return today, today + timedelta(days=FUTURE_DAYS)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
@router.get("/export/csv")
async def export_csv(
    scope: str = Query("future"),
    team: str  = Query(""),
):
    start, end = _get_range(scope)
    dates      = date_range(start, end)
    rows      = get_roster_for_range(start.isoformat(), end.isoformat(), team)
    overrides = get_overrides_for_range(start.isoformat(), end.isoformat())
    agents, dates = build_matrix(rows, dates, overrides)

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Row 1 – day names
    writer.writerow([""] * len(_IDENTITY_HEADERS) + [d.strftime("%a") for d in dates])
    # Row 2 – headers
    writer.writerow(_IDENTITY_HEADERS + [d.strftime("%d-%b-%Y") for d in dates])

    for i, agent in enumerate(agents, start=1):
        identity = [
            i,
            agent.get("login_id", ""),
            agent.get("win_id", ""),
            agent["full_name"],
            agent.get("team_name", ""),
            agent.get("role", ""),
            agent.get("l1_manager", ""),
        ]
        cells = [_lbl(agent["schedule"].get(d.isoformat(), {"label":"OFF"})) for d in dates]
        writer.writerow(identity + cells)

    buf.seek(0)
    fname = f"transport_roster_{scope}_{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
@router.get("/export/excel")
async def export_excel(
    scope: str = Query("future"),
    team: str  = Query(""),
):
    try:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import (  # noqa: PLC0415
            PatternFill, Font, Alignment, Border, Side,
        )
    except ImportError:
        return StreamingResponse(
            iter([b"openpyxl not installed"]),
            media_type="text/plain",
        )

    start, end = _get_range(scope)
    dates      = date_range(start, end)
    rows      = get_roster_for_range(start.isoformat(), end.isoformat(), team)
    overrides = get_overrides_for_range(start.isoformat(), end.isoformat())
    agents, dates = build_matrix(rows, dates, overrides)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transport Roster"

    # -- Style definitions --------------------------------------------------
    hdr_fill   = PatternFill("solid", fgColor="0053E2")   # Walmart blue
    day_fill   = PatternFill("solid", fgColor="FFC220")   # Spark yellow
    off_fill   = PatternFill("solid", fgColor="F3F4F6")
    on_fill    = PatternFill("solid", fgColor="D1FAE5")   # light green
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    day_font   = Font(bold=True, color="1F2937", size=9)
    data_font  = Font(size=9)
    off_font   = Font(size=9, color="9CA3AF")
    thin       = Side(style="thin", color="D1D5DB")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    n_id  = len(_IDENTITY_HEADERS)
    n_col = n_id + len(dates)

    # Row 1 – day names (date columns only)
    for col_idx, d in enumerate(dates, start=n_id + 1):
        cell = ws.cell(row=1, column=col_idx, value=d.strftime("%a"))
        cell.fill      = day_fill
        cell.font      = day_font
        cell.alignment = center
        cell.border    = border
    # Blank identity cells row 1
    for col_idx in range(1, n_id + 1):
        ws.cell(row=1, column=col_idx).fill   = hdr_fill
        ws.cell(row=1, column=col_idx).border = border

    # Row 2 – column headers
    all_headers = _IDENTITY_HEADERS + [d.strftime("%d-%b-%Y") for d in dates]
    for col_idx, h in enumerate(all_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border

    # Data rows
    for row_idx, agent in enumerate(agents, start=3):
        i = row_idx - 2
        identity = [
            i, agent.get("login_id",""), agent.get("win_id",""), agent["full_name"],
            agent.get("team_name",""), agent.get("role",""), agent.get("l1_manager",""),
        ]
        for col_idx, val in enumerate(identity, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.alignment = left_align
            cell.border    = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

        for col_idx, d in enumerate(dates, start=n_id + 1):
            cell_data = agent["schedule"].get(d.isoformat(), {"label":"OFF","is_override":False})
            label = _lbl(cell_data)
            is_ovr = cell_data.get("is_override", False) if isinstance(cell_data, dict) else False
            cell  = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.alignment = center
            cell.border    = border
            if label == "OFF":
                cell.fill = off_fill
                cell.font = off_font
            elif is_ovr:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")   # override amber
                cell.font = Font(size=9, color="92400E")
            else:
                cell.fill = on_fill
                cell.font = Font(size=9, color="166534")

    # Column widths
    id_widths = [5, 11, 13, 24, 28, 18, 20]
    for i, w in enumerate(id_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for col_idx in range(n_id + 1, n_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "H3"  # freeze identity cols + both header rows

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"transport_roster_{scope}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
@router.get("/export/pdf")
async def export_pdf(
    scope: str = Query("future"),
    team: str  = Query(""),
):
    try:
        from fpdf import FPDF  # noqa: PLC0415
    except ImportError:
        return StreamingResponse(
            iter([b"fpdf2 not installed"]),
            media_type="text/plain",
        )

    start, end = _get_range(scope)
    dates      = date_range(start, end)
    rows      = get_roster_for_range(start.isoformat(), end.isoformat(), team)
    overrides = get_overrides_for_range(start.isoformat(), end.isoformat())
    agents, dates = build_matrix(rows, dates, overrides)

    # Split dates into pages of 14 date-cols each (landscape A4 fits ~14)
    DATES_PER_PAGE = 14
    chunks = [dates[i:i + DATES_PER_PAGE] for i in range(0, len(dates), DATES_PER_PAGE)]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)

    title_label = f"CES IND Transport Roster - {scope.title()} Schedule"
    label = team or "All Teams"

    for chunk_dates in chunks:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_fill_color(0, 83, 226)     # Walmart blue
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 8, title_label, ln=True, fill=True, align="C")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 5, f"Team: {label}  |  Generated: {date.today().isoformat()}", ln=True, align="C")
        pdf.ln(2)

        # Column widths
        id_w   = [6, 18, 18, 36, 32, 20, 30]   # identity col widths mm
        date_w = 22                              # each date col mm

        def draw_header(chunk: list[date]) -> None:
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_fill_color(0, 83, 226)
            pdf.set_text_color(255, 255, 255)
            for w, h in zip(id_w, _IDENTITY_HEADERS):
                pdf.cell(w, 6, h, border=1, align="C", fill=True)
            for d in chunk:
                pdf.cell(date_w, 6, d.strftime("%d %b"), border=1, align="C", fill=True)
            pdf.ln()

        draw_header(chunk_dates)

        for i, agent in enumerate(agents):
            pdf.set_font("Helvetica", "", 6.5)
            pdf.set_text_color(30, 30, 30)
            fill_bg = (249, 250, 251) if i % 2 == 0 else (255, 255, 255)
            pdf.set_fill_color(*fill_bg)

            identity = [
                str(i + 1), agent.get("login_id",""), agent.get("win_id",""),
                agent["full_name"][:22], agent.get("team_name","")[:20],
                agent.get("role","")[:12], agent.get("l1_manager","")[:18],
            ]
            for w, val in zip(id_w, identity):
                pdf.cell(w, 5, val, border=1, align="L", fill=True)

            for d in chunk_dates:
                label_cell = _lbl(agent["schedule"].get(d.isoformat(), {"label":"OFF"}))
                if label_cell == "OFF":
                    pdf.set_fill_color(243, 244, 246)
                    pdf.set_text_color(156, 163, 175)
                else:
                    pdf.set_fill_color(209, 250, 229)
                    pdf.set_text_color(22, 101, 52)
                pdf.cell(date_w, 5, label_cell, border=1, align="C", fill=True)
                pdf.set_fill_color(*fill_bg)
                pdf.set_text_color(30, 30, 30)
            pdf.ln()

    buf = io.BytesIO(pdf.output())
    fname = f"transport_roster_{scope}_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
