"""
routes/admin.py  - Admin-only endpoints:
    GET  /admin/overrides         override editor page
    GET  /api/admin/agents        JSON: agents list (filtered by l1_manager)
    POST /api/admin/override      save / update one override
    DELETE /api/admin/override    remove one override
    GET  /admin/adhoc             ad-hoc roster page
    POST /admin/adhoc/preview     HTMX: preview ad-hoc matrix
    GET  /admin/adhoc/export      download ad-hoc roster (csv/excel/pdf)
"""
import sys
import io
import csv
from pathlib import Path
from datetime import date, timedelta, datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

from fastapi import APIRouter, Request, Query, Form, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from config import APP_DIR
from auth import is_admin
from db import (
    get_l1_managers, get_agents_for_manager, get_all_agents,
    upsert_override, delete_override,
    get_roster_for_range, get_overrides_for_range,
    get_master_stats,
)
from master import load_excel_into_master
from matrix import date_range, build_matrix

router    = APIRouter()
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))

_SHIFT_PRESETS = {
    "07:00 AM - 04:00 PM": "07:00 AM - 04:00 PM",
    "08:00 AM - 05:00 PM": "08:00 AM - 05:00 PM",
    "09:00 AM - 06:00 PM": "09:00 AM - 06:00 PM",
    "10:00 AM - 07:00 PM": "10:00 AM - 07:00 PM",
    "11:00 AM - 08:00 PM": "11:00 AM - 08:00 PM",
    "12:00 PM - 09:00 PM": "12:00 PM - 09:00 PM",
    "01:30 PM - 10:30 PM": "01:30 PM - 10:30 PM",
    "Half Day AM":          "Half Day AM",
    "Half Day PM":          "Half Day PM",
    "OFF":                  "OFF",
}


def _require_admin(request: Request):
    return is_admin(request)


# ---------------------------------------------------------------------------
# Master Data Upload
# ---------------------------------------------------------------------------
@router.get("/admin/master", response_class=HTMLResponse)
async def master_page(request: Request):
    if not _require_admin(request):
        return RedirectResponse("/login?next=/admin/master")
    stats = get_master_stats()
    return templates.TemplateResponse("admin/master_upload.html", {
        "request": request,
        "stats":   stats,
    })


@router.post("/admin/master/upload", response_class=HTMLResponse)
async def master_upload(
    request: Request,
    files: list[UploadFile] = File(...),
):
    if not _require_admin(request):
        return HTMLResponse("<p class='text-red-500'>Unauthorised</p>", status_code=403)

    import tempfile, shutil
    results = []
    for upload in files:
        if not upload.filename:
            continue
        # Save to a temp file so openpyxl can read it
        suffix = Path(upload.filename).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            shutil.copyfileobj(upload.file, tmp)
            tmp_path = tmp.name
        label = Path(upload.filename).stem
        result = load_excel_into_master(tmp_path, label)
        result["file"] = upload.filename   # use original filename in UI
        Path(tmp_path).unlink(missing_ok=True)
        results.append(result)

    stats = get_master_stats()
    # Return an HTMX partial
    rows_html = "".join(
        f"""<tr class='border-b'>
            <td class='py-2 px-3 font-mono text-sm'>{r['file']}</td>
            <td class='py-2 px-3 text-center'>{r['parsed']}</td>
            <td class='py-2 px-3 text-center text-green-700 font-bold'>{r['upserted']}</td>
            <td class='py-2 px-3 text-center text-{'red' if r['error'] else 'gray'}-500'>
                {' ' + r['error'] if r['error'] else ' OK'}
            </td>
        </tr>"""
        for r in results
    )
    return HTMLResponse(f"""
        <div class='mt-4 p-4 bg-green-50 border border-green-200 rounded-lg'>
            <p class='font-bold text-green-800 mb-3'> Upload complete — Master DB now has
               <span class='text-blue-700'>{stats['total']} agents</span>
               ({stats['active']} active · {stats['lob_count']} LOBs)</p>
            <table class='w-full text-sm border border-gray-200 rounded'>
                <thead class='bg-gray-50'><tr>
                    <th class='py-2 px-3 text-left'>File</th>
                    <th class='py-2 px-3 text-center'>Parsed</th>
                    <th class='py-2 px-3 text-center'>Saved</th>
                    <th class='py-2 px-3 text-center'>Status</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
            <p class='text-xs text-gray-500 mt-2'>Reload any roster tab to see updated names, teams & managers.</p>
        </div>
    """)


@router.post("/admin/master/reload-defaults", response_class=HTMLResponse)
async def master_reload_defaults(request: Request):
    """Re-load from the OneDrive HeadCount files (local machine only)."""
    if not _require_admin(request):
        return HTMLResponse("<p class='text-red-500'>Unauthorised</p>", status_code=403)
    from master import _DEFAULT_FILES, load_excel_into_master
    results = []
    for file_path, label in _DEFAULT_FILES:
        result = load_excel_into_master(file_path, label)
        results.append(result)
    stats = get_master_stats()
    rows_html = "".join(
        f"""<tr class='border-b'>
            <td class='py-2 px-3 font-mono text-sm'>{Path(r['file']).name if '\\' in r['file'] or '/' in r['file'] else r['file']}</td>
            <td class='py-2 px-3 text-center'>{r['parsed']}</td>
            <td class='py-2 px-3 text-center text-green-700 font-bold'>{r['upserted']}</td>
            <td class='py-2 px-3 text-center text-{'red' if r['error'] else 'gray'}-500'>
                {' ' + r['error'] if r['error'] else ' OK'}
            </td>
        </tr>"""
        for r in results
    )
    return HTMLResponse(f"""
        <div class='mt-4 p-4 bg-blue-50 border border-blue-200 rounded-lg'>
            <p class='font-bold text-blue-800 mb-3'> Reload complete — Master DB now has
               <span class='text-blue-700'>{stats['total']} agents</span>
               ({stats['active']} active · {stats['lob_count']} LOBs)</p>
            <table class='w-full text-sm border border-gray-200 rounded'>
                <thead class='bg-gray-50'><tr>
                    <th class='py-2 px-3 text-left'>File</th>
                    <th class='py-2 px-3 text-center'>Parsed</th>
                    <th class='py-2 px-3 text-center'>Saved</th>
                    <th class='py-2 px-3 text-center'>Status</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>
    """)


# ---------------------------------------------------------------------------
# Override Editor
# ---------------------------------------------------------------------------
@router.get("/admin/overrides", response_class=HTMLResponse)
async def override_page(request: Request, l1: str = Query("")):
    if not _require_admin(request):
        return RedirectResponse("/login?next=/admin/overrides")
    managers = get_l1_managers()
    agents   = get_agents_for_manager(l1) if l1 else []
    today    = date.today()
    end      = today + timedelta(days=14)
    overrides = get_overrides_for_range(today.isoformat(), end.isoformat())

    # Build a mini matrix for selected agents if any
    matrix_agents: list[dict] = []
    dates: list[date] = []
    if agents:
        vcc_ids = [a["vcc_id"] for a in agents]
        rows = get_roster_for_range(today.isoformat(), end.isoformat())
        rows = [r for r in rows if r["vcc_id"] in vcc_ids]
        dates = date_range(today, end)
        matrix_agents, dates = build_matrix(rows, dates, overrides)

    return templates.TemplateResponse("admin/overrides.html", {
        "request":        request,
        "managers":       managers,
        "selected_l1":    l1,
        "agents":         agents,
        "matrix_agents":  matrix_agents,
        "dates":          dates,
        "presets":        list(_SHIFT_PRESETS.keys()),
        "today":          today.isoformat(),
    })


@router.get("/api/admin/agents", response_class=JSONResponse)
async def api_agents(request: Request, l1: str = Query("")):
    if not _require_admin(request):
        return JSONResponse({"error": "Unauthorised"}, status_code=403)
    agents = get_agents_for_manager(l1) if l1 else get_all_agents()
    return JSONResponse(agents)


@router.post("/api/admin/override")
async def save_override(
    request: Request,
    vcc_id: str        = Form(...),
    full_name: str     = Form(""),
    schedule_date: str = Form(...),
    shift_label: str   = Form(...),
    note: str          = Form(""),
):
    if not _require_admin(request):
        return JSONResponse({"error": "Unauthorised"}, status_code=403)
    upsert_override(vcc_id, full_name, schedule_date, shift_label, note)
    return JSONResponse({"success": True, "message": f"Override saved for {vcc_id} on {schedule_date}"})


@router.post("/api/admin/override/delete")
async def remove_override(
    request: Request,
    vcc_id: str        = Form(...),
    schedule_date: str = Form(...),
):
    if not _require_admin(request):
        return JSONResponse({"error": "Unauthorised"}, status_code=403)
    delete_override(vcc_id, schedule_date)
    return JSONResponse({"success": True})


# ---------------------------------------------------------------------------
# Ad-Hoc Roster
# ---------------------------------------------------------------------------
@router.get("/admin/adhoc", response_class=HTMLResponse)
async def adhoc_page(request: Request):
    if not _require_admin(request):
        return RedirectResponse("/login?next=/admin/adhoc")
    return templates.TemplateResponse("admin/adhoc.html", {
        "request":  request,
        "managers": get_l1_managers(),
        "agents":   get_all_agents(),
        "presets":  list(_SHIFT_PRESETS.keys()),
        "today":    date.today().isoformat(),
    })


def _build_adhoc_matrix(vcc_ids: list[str], start: date, end: date,
                         shift_label: str, week_off_days: list[int]) -> tuple[list[dict], list[date]]:
    """Build an in-memory ad-hoc schedule without touching the DB."""
    all_agents = get_all_agents()
    agents_map = {a["vcc_id"]: a for a in all_agents if a["vcc_id"] in vcc_ids}
    dates = date_range(start, end)
    result: list[dict] = []
    for vcc, agent in sorted(agents_map.items(), key=lambda x: x[1]["full_name"].lower()):
        schedule = {}
        for d in dates:
            ds = d.isoformat()
            schedule[ds] = {
                "label":       "OFF" if d.weekday() in week_off_days else shift_label,
                "shift_type":  "FULL",
                "is_override": False,
                "note":        "",
            }
        entry = dict(agent)
        entry["schedule"] = schedule
        result.append(entry)
    return result, dates


@router.post("/admin/adhoc/preview", response_class=HTMLResponse)
async def adhoc_preview(
    request: Request,
    vcc_ids: list[str]      = Form(...),
    start_date: str         = Form(...),
    end_date: str           = Form(...),
    shift_label: str        = Form(...),
    week_off: list[int]     = Form([]),
):
    if not _require_admin(request):
        return HTMLResponse("<p class='text-red-500'>Unauthorised</p>", status_code=403)
    start = date.fromisoformat(start_date)
    end   = date.fromisoformat(end_date)
    agents, dates = _build_adhoc_matrix(vcc_ids, start, end, shift_label, week_off)
    return templates.TemplateResponse("partials/roster_table.html", {
        "request":       request,
        "agents":        agents,
        "dates":         dates,
        "teams":         [],
        "selected_team": "",
        "tab":           "adhoc",
        "empty_message": "No agents selected.",
        "is_admin":      True,
    })


@router.get("/admin/adhoc/export")
async def adhoc_export(
    request: Request,
    vcc_ids: str  = Query(""),
    start_date: str = Query(""),
    end_date: str   = Query(""),
    shift_label: str = Query(""),
    week_off: str    = Query(""),
    fmt: str         = Query("excel"),
):
    if not _require_admin(request):
        return JSONResponse({"error": "Unauthorised"}, status_code=403)

    ids   = [v for v in vcc_ids.split(",") if v]
    start = date.fromisoformat(start_date) if start_date else date.today()
    end   = date.fromisoformat(end_date) if end_date else start + timedelta(days=30)
    wo    = [int(x) for x in week_off.split(",") if x.isdigit()]
    agents, dates = _build_adhoc_matrix(ids, start, end, shift_label, wo)

    if fmt == "csv":
        return _adhoc_csv(agents, dates)
    if fmt == "pdf":
        return _adhoc_pdf(agents, dates, shift_label)
    return _adhoc_excel(agents, dates)


# -- Export helpers ---------------------------------------------------------
_ID_HEADERS = ["S.No", "User ID", "WIN ID", "Full Name", "Team", "Role", "L1 Manager"]


def _identity(i: int, a: dict) -> list:
    return [i, a.get("login_id",""), a.get("win_id",""), a["full_name"],
            a.get("team_name",""), a.get("role",""), a.get("l1_manager","")]


def _adhoc_csv(agents: list[dict], dates: list[date]) -> StreamingResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([""] * len(_ID_HEADERS) + [d.strftime("%a") for d in dates])
    w.writerow(_ID_HEADERS + [d.strftime("%d-%b-%Y") for d in dates])
    for i, a in enumerate(agents, 1):
        w.writerow(_identity(i, a) + [a["schedule"][d.isoformat()]["label"] for d in dates])
    buf.seek(0)
    fname = f"adhoc_roster_{date.today().isoformat()}.csv"
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


def _adhoc_excel(agents: list[dict], dates: list[date]) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ad-Hoc Roster"
    hdr_fill  = PatternFill("solid", fgColor="0053E2")
    day_fill  = PatternFill("solid", fgColor="FFC220")
    on_fill   = PatternFill("solid", fgColor="D1FAE5")
    off_fill  = PatternFill("solid", fgColor="F3F4F6")
    thin      = Side(style="thin", color="D1D5DB")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    n_id = len(_ID_HEADERS)

    for ci, d in enumerate(dates, n_id + 1):
        c = ws.cell(1, ci, d.strftime("%a"))
        c.fill = day_fill; c.font = Font(bold=True, size=9); c.alignment = center; c.border = border
    for ci in range(1, n_id + 1):
        ws.cell(1, ci).fill = hdr_fill; ws.cell(1, ci).border = border

    for ci, h in enumerate(_ID_HEADERS + [d.strftime("%d-%b-%Y") for d in dates], 1):
        c = ws.cell(2, ci, h)
        c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center; c.border = border

    for ri, a in enumerate(agents, 3):
        for ci, val in enumerate(_identity(ri - 2, a), 1):
            c = ws.cell(ri, ci, val); c.border = border
            c.font = Font(size=9); c.alignment = Alignment(horizontal="left", vertical="center")
        for ci, d in enumerate(dates, n_id + 1):
            lbl = a["schedule"][d.isoformat()]["label"]
            c = ws.cell(ri, ci, lbl); c.alignment = center; c.border = border
            c.fill = off_fill if lbl == "OFF" else on_fill
            c.font = Font(size=9, color="9CA3AF" if lbl == "OFF" else "166534")

    for i, w in enumerate([5, 11, 13, 24, 28, 18, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for ci in range(n_id + 1, n_id + len(dates) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22
    ws.freeze_panes = "H3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"adhoc_roster_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


def _adhoc_pdf(agents: list[dict], dates: list[date], shift_label: str) -> StreamingResponse:
    from fpdf import FPDF
    DATES_PER_PAGE = 14
    chunks = [dates[i:i+DATES_PER_PAGE] for i in range(0, len(dates), DATES_PER_PAGE)]
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    id_w = [6, 18, 18, 36, 32, 20, 30]

    for chunk in chunks:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_fill_color(0, 83, 226); pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 8, "CES IND Transport - Ad-Hoc Roster", ln=True, fill=True, align="C")
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 5, f"Shift: {shift_label}  |  Generated: {date.today().isoformat()}", ln=True, align="C")
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 7); pdf.set_fill_color(0, 83, 226); pdf.set_text_color(255, 255, 255)
        for w, h in zip(id_w, _ID_HEADERS):
            pdf.cell(w, 6, h, border=1, align="C", fill=True)
        for d in chunk:
            pdf.cell(22, 6, d.strftime("%d %b"), border=1, align="C", fill=True)
        pdf.ln()
        for i, a in enumerate(agents):
            pdf.set_font("Helvetica", "", 6.5); pdf.set_text_color(30, 30, 30)
            bg = (249, 250, 251) if i % 2 == 0 else (255, 255, 255)
            pdf.set_fill_color(*bg)
            for w, val in zip(id_w, _identity(i+1, a)):
                pdf.cell(w, 5, str(val)[:20], border=1, align="L", fill=True)
            for d in chunk:
                lbl = a["schedule"][d.isoformat()]["label"]
                if lbl == "OFF":
                    pdf.set_fill_color(243, 244, 246); pdf.set_text_color(156, 163, 175)
                else:
                    pdf.set_fill_color(209, 250, 229); pdf.set_text_color(22, 101, 52)
                pdf.cell(22, 5, lbl, border=1, align="C", fill=True)
                pdf.set_fill_color(*bg); pdf.set_text_color(30, 30, 30)
            pdf.ln()

    buf = io.BytesIO(pdf.output())
    fname = f"adhoc_roster_{date.today().isoformat()}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
