"""
master.py - Parse HeadCount Excel files and load into agent_master table.

HeadCount_Base column mapping (0-indexed):
  [01] User ID      -> login_id  (key - stored lowercase)
  [02] WIN ID       -> win_id
  [09] First Name   -> first_name
  [10] Last Name    -> last_name
  [11] Associate Name (WD) -> full_name
  [12] Role         -> role
  [13] Employee Current Status -> status
  [15] LOB          -> lob  (used as team_name in the app)
  [29] Location     -> location
  [52] Team Lead Name     -> l1_manager
  [53] Team Manager Name  -> l2_manager
  [54] Ops Manager Name   -> ops_manager
"""
import logging
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from db import upsert_master, get_master_stats

logger = logging.getLogger(__name__)

# Columns we need (0-indexed positions in HeadCount_Base sheet)
_COL = {
    "login_id":   1,
    "win_id":     2,
    "first_name": 9,
    "last_name":  10,
    "full_name":  11,
    "role":       12,
    "status":     13,
    "lob":        15,
    "location":   29,
    "l1_manager": 52,
    "l2_manager": 53,
    "ops_manager": 54,
}

_SHEET = "HeadCount_Base"


def _clean(val) -> str:
    """Normalise a cell value to a clean string."""
    if val is None:
        return ""
    if isinstance(val, float):
        # WIN IDs stored as floats like 231505622.0
        return str(int(val))
    return str(val).strip()


def parse_headcount_excel(file_path: str | Path, source_label: str = "") -> list[dict]:
    """
    Parse one HeadCount Excel file.
    Returns a list of dicts ready for upsert_master().
    Skips rows with a blank/missing login_id.
    """
    path = Path(file_path)
    if not path.exists():
        logger.warning("Master file not found: %s", path)
        return []

    label = source_label or path.stem
    loaded_at = datetime.now(timezone.utc).isoformat()

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Cannot open %s: %s", path.name, exc)
        return []

    if _SHEET not in wb.sheetnames:
        logger.error("Sheet '%s' not found in %s. Sheets: %s",
                     _SHEET, path.name, wb.sheetnames)
        wb.close()
        return []

    ws = wb[_SHEET]
    records: list[dict] = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip if row is shorter than expected or login_id is blank
        if not row or len(row) <= _COL["login_id"]:
            skipped += 1
            continue

        login_id = _clean(row[_COL["login_id"]]).lower()
        if not login_id or login_id == "user id":
            skipped += 1
            continue

        records.append({
            "login_id":   login_id,
            "win_id":     _clean(row[_COL["win_id"]]) if len(row) > _COL["win_id"] else "",
            "first_name": _clean(row[_COL["first_name"]]) if len(row) > _COL["first_name"] else "",
            "last_name":  _clean(row[_COL["last_name"]]) if len(row) > _COL["last_name"] else "",
            "full_name":  _clean(row[_COL["full_name"]]) if len(row) > _COL["full_name"] else "",
            "role":       _clean(row[_COL["role"]]) if len(row) > _COL["role"] else "",
            "status":     _clean(row[_COL["status"]]) if len(row) > _COL["status"] else "",
            "lob":        _clean(row[_COL["lob"]]) if len(row) > _COL["lob"] else "",
            "location":   _clean(row[_COL["location"]]) if len(row) > _COL["location"] else "",
            "l1_manager": _clean(row[_COL["l1_manager"]]) if len(row) > _COL["l1_manager"] else "",
            "l2_manager": _clean(row[_COL["l2_manager"]]) if len(row) > _COL["l2_manager"] else "",
            "ops_manager": _clean(row[_COL["ops_manager"]]) if len(row) > _COL["ops_manager"] else "",
            "source_file": label,
            "loaded_at":  loaded_at,
        })

    wb.close()
    logger.info("Parsed %d records (%d skipped) from %s", len(records), skipped, path.name)
    return records


def load_excel_into_master(file_path: str | Path, source_label: str = "") -> dict:
    """
    Parse one HeadCount Excel and upsert into agent_master.
    Returns a summary dict: {file, parsed, upserted, error}.
    """
    path = Path(file_path)
    try:
        records = parse_headcount_excel(path, source_label or path.stem)
        if not records:
            return {"file": path.name, "parsed": 0, "upserted": 0,
                    "error": f"No records parsed — check '{_SHEET}' sheet exists."}
        count = upsert_master(records)
        logger.info("Upserted %d master records from %s", count, path.name)
        return {"file": path.name, "parsed": len(records), "upserted": count, "error": ""}
    except Exception as exc:
        logger.error("Failed loading master from %s: %s", path.name, exc)
        return {"file": path.name, "parsed": 0, "upserted": 0, "error": str(exc)}


# ---------------------------------------------------------------------------
# Auto-loader — called at startup if agent_master is empty
# ---------------------------------------------------------------------------
_DEFAULT_FILES = [
    (
        r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - "
        r"CareOps_Centralized_Report\12.HeadCount\Head_Count_MAA.xlsx",
        "MAA",
    ),
    (
        r"C:\Users\V0M06TT\OneDrive - Walmart Inc\Shared Documents - "
        r"CareOps_Centralized_Report\12.HeadCount\Head_Count_BLR.xlsx",
        "BLR",
    ),
]


def auto_load_master_if_empty() -> dict:
    """
    On startup: if agent_master table is empty, load from the known
    OneDrive HeadCount files automatically.
    Returns a summary of what happened.
    """
    stats = get_master_stats()
    if stats["total"] > 0:
        logger.info(
            "Agent master already has %d records — skipping auto-load.", stats["total"]
        )
        return {"skipped": True, "total": stats["total"]}

    logger.info("Agent master is empty — attempting auto-load from HeadCount files.")
    results = []
    for file_path, label in _DEFAULT_FILES:
        result = load_excel_into_master(file_path, label)
        results.append(result)
        if result["error"]:
            logger.warning("Auto-load issue for %s: %s", label, result["error"])

    final_stats = get_master_stats()
    logger.info(
        "Auto-load complete: %d total master records loaded.", final_stats["total"]
    )
    return {"skipped": False, "results": results, "total": final_stats["total"]}
